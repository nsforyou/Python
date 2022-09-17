# 课程名称:Python基础课程
# 创建日期:2022/9/18 2:07
# 作   者: 十八般武艺
# 版   本: Python 3.10.7

"""
授课内容:

"""
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

# Getting sheets from the workbook

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

mysheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

#sheet3 = wb.get_sheet_by_name('sheet3')
#sheet4 = wb['mySheet'])

# 从活跃工作表获取单元格

ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)